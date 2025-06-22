# Librerías necesarias
import requests  # Para hacer solicitudes HTTP (API)
from openpyxl import load_workbook, Workbook  # Para leer/escribir archivos Excel
from datetime import datetime, timedelta  # Para trabajar con fechas
import smtplib  # Para enviar correos electrónicos
from email.mime.multipart import MIMEMultipart  # Para construir mensajes con adjuntos
from email.mime.base import MIMEBase  # Para adjuntar archivos
from email import encoders  # Para codificar archivos adjuntos
import logging  # Para registrar errores y eventos

# ----------------------------
# 1. Clase para conectarse a Odoo y obtener ventas
# ----------------------------
class OdooSalesData:
    def __init__(self, api_url):
        self.api_url = api_url  # Guarda la URL base del servidor Odoo
    
    def get_daily_sales(self, date):
        """Obtiene ventas de una fecha específica desde Odoo"""
        response = requests.get(
            f"{self.api_url}/export_daily_sales",  # Endpoint de ventas diarias
            params={"date": date.strftime("%Y-%m-%d")}  # Parámetro de fecha en formato YYYY-MM-DD
        )
        return response.json()['sales']  # Devuelve sólo la lista de ventas

# ----------------------------
# 2. Clase para consultar cotización del dólar desde una API externa
# ----------------------------
class DollarAPI:
    def __init__(self, api_endpoint):
        self.endpoint = api_endpoint  # URL de la API del dólar
    
    def get_historical_rate(self, date):
        """Consulta la cotización del dólar oficial para una fecha"""
        try:
            response = requests.get(
                self.endpoint,
                params={"date": date.strftime("%Y-%m-%d")}
            )
            return response.json()['oficial']['value_avg']  # Devuelve valor promedio
        except Exception as e:
            logging.error(f"Error API Dólar: {e}")  # Registra error si la API falla
            return None

# ----------------------------
# 3. Clase para generar o modificar un archivo Excel con los datos
# ----------------------------
class SalesReportExcel:
    TEMPLATE_PATH = "/templates/report_template.xlsx"  # Ruta a una plantilla (no usada en este código)

    def __init__(self, file_path):
        self.file_path = file_path  # Ruta donde se guardará el Excel
        try:
            self.workbook = load_workbook(file_path)  # Intenta cargar archivo existente
        except FileNotFoundError:
            self.workbook = Workbook()  # Crea uno nuevo si no existe
            self.workbook.remove(self.workbook.active)  # Elimina hoja vacía por defecto

    def add_daily_sheet(self, data, usd_rate):
        """Crea una nueva hoja en el Excel con datos de ventas"""
        sheet_name = datetime.now().strftime("%Y-%m-%d")  # Nombre de hoja = fecha actual
        ws = self.workbook.create_sheet(sheet_name)  # Crea nueva hoja

        # Agrega encabezados de columnas
        ws.append(["Cliente", "Productos", "Importe USD", "Fecha", "Importe ARS", "Promedio Diario", "Acumulado"])

        # Agrega una fila por cada venta
        for sale in data:
            ws.append([
                sale['client'],
                sale['products'],
                sale['amount'],
                sale['date'],
                f"=C{ws.max_row}*{usd_rate}",  # Cálculo en ARS como fórmula Excel
                "",  # Promedio y acumulado se llenan luego
                ""
            ])

        self._calculate_financials(ws, usd_rate)  # Aplica fórmulas de promedio y acumulado

    def _calculate_financials(self, worksheet, usd_rate):
        """Agrega fórmulas para calcular totales y bordes"""
        days_count = len([s for s in self.workbook.sheetnames if s != worksheet.title])  # Cuenta hojas anteriores

        last_row = worksheet.max_row  # Última fila con datos

        worksheet[f"F2"] = f"=E2/{days_count}"  # Fórmula: promedio diario
        worksheet[f"G2"] = f"=SUM(E2:E{last_row})"  # Fórmula: acumulado en ARS

        # Aplica bordes gruesos a todas las celdas con datos
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = self._get_thick_border()

    def save(self):
        self.workbook.save(self.file_path)  # Guarda el Excel en disco

    @staticmethod
    def _get_thick_border():
        from openpyxl.styles import Border, Side
        return Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

# ----------------------------
# 4. Clase para enviar el Excel por correo electrónico
# ----------------------------
class EmailNotifier:
    def __init__(self, smtp_config):
        self.smtp_config = smtp_config  # Diccionario con datos del servidor SMTP

    def send_report(self, recipient, attachment_path):
        msg = MIMEMultipart()  # Crea el mensaje de email
        msg['Subject'] = f"Reporte Ventas {datetime.today().strftime('%Y-%m-%d')}"  # Asunto

        # Adjunta el archivo Excel al correo
        with open(attachment_path, "rb") as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{attachment_path.name}"')
            msg.attach(part)

        # Envia el correo usando SMTP (con TLS)
        with smtplib.SMTP(self.smtp_config['host'], self.smtp_config['port']) as server:
            server.starttls()  # Inicia modo seguro
            server.login(self.smtp_config['user'], self.smtp_config['password'])  # Inicia sesión
            server.sendmail(self.smtp_config['user'], recipient, msg.as_string())  # Envía el correo

# ----------------------------
# 5. Clase principal que orquesta todo el proceso
# ----------------------------
class DailySalesReport:
    def __init__(self):
        # Inicializa los componentes necesarios
        self.odoo = OdooSalesData("http://odoo-instance.com")
        self.dollar_api = DollarAPI("https://api.bluelytics.com.ar/v2/evolution.json")
        self.excel = SalesReportExcel("/reports/daily_sales.xlsx")
        self.notifier = EmailNotifier({
            'host': 'smtp.gmail.com',
            'port': 587,
            'user': 'bot@empresa.com',
            'password': 'password'
        })

    def generate_report(self):
        try:
            # Obtiene las fechas de hoy y ayer
            today = datetime.now()
            yesterday = today - timedelta(days=1)

            # 1. Obtiene datos de ventas y cotización del dólar
            sales_data = self.odoo.get_daily_sales(today)
            usd_rate = self.dollar_api.get_historical_rate(yesterday)

            if not sales_data:
                logging.warning("No hay ventas para reportar hoy")
                return  # Termina si no hay datos

            # 2. Crea Excel con los datos
            self.excel.add_daily_sheet(sales_data, usd_rate)
            self.excel.save()

            # 3. Envía el Excel por email
            self.notifier.send_report('finanzas@empresa.com', self.excel.file_path)

        except Exception as e:
            logging.error(f"Error generando reporte: {e}")  # Registra error
            raise

# ----------------------------
# Ejecuta el programa si es archivo principal
# ----------------------------
if __name__ == "__main__":
    report = DailySalesReport()  # Crea una instancia del reporte
    report.generate_report()     # Ejecuta todo el proceso